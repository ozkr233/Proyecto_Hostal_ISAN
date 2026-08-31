"""Lectura de los libros Excel a registros crudos (dicts de texto).

Aqui no se toca la base de datos ni se valida nada de negocio: la unica
responsabilidad es entender el layout de cada hoja y devolver filas planas.
Eso permite ejecutar el ETL en modo --dry-run sin Postgres.

La regla central: **el bloque de ingreso y el de salida de una hoja diaria son
dos listas independientes puestas lado a lado, no la misma persona**. En
'01- JULIO' la fila 14 no tiene ingreso y si registra la salida de ALAN
CALDERON. Por eso cada fila del Excel puede producir cero, uno o dos registros.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from . import config
from . import normalizar as nz


@dataclass
class Mapeo:
    """Columnas (indice 0) de cada campo canonico, por bloque."""
    ingreso: dict[str, list[int]] = field(default_factory=dict)
    salida: dict[str, list[int]] = field(default_factory=dict)
    inicio_salida: int | None = None

    def valor(self, fila: tuple, bloque: str, campo: str):
        """Primer valor no vacio entre las columnas asignadas al campo.

        Varios encabezados se repiten dentro de una misma hoja ('TURNO'
        aparece dos veces en el bloque de ingreso de ISAM, y una de las dos
        columnas viene siempre vacia). Tomar el primer valor util evita tener
        que adivinar cual de las dos es la buena.
        """
        cols = (self.ingreso if bloque == "INGRESO" else self.salida).get(campo, [])
        for c in cols:
            if c < len(fila):
                v = fila[c]
                if v is not None and str(v).strip() not in ("", "-"):
                    return v
        return None


def _mapear(encabezados: list[str | None], sinonimos: dict, desde: int, hasta: int) -> dict[str, list[int]]:
    """Asigna columnas del rango [desde, hasta) a campos canonicos."""
    mapeo: dict[str, list[int]] = {}
    for idx in range(desde, min(hasta, len(encabezados))):
        h = encabezados[idx]
        if not h:
            continue
        for campo, alternativas in sinonimos.items():
            if h in alternativas:
                mapeo.setdefault(campo, []).append(idx)
                break
    return mapeo


def detectar_mapeo_diario(encabezados_crudos: list) -> Mapeo:
    """Deduce el layout de una hoja diaria leyendo su fila de encabezado.

    El corte entre bloques se encuentra asi: se ubica la columna del NOMBRE de
    ingreso y luego la primera columna de fecha que venga despues. Esa es la
    fecha de salida, y ahi empieza el bloque de salida.

    Con esta regla los dias 5 y 6 -- donde OBSERVACIONES se cuela antes de la
    fecha de salida y corre todo una columna -- se resuelven solos, sin un caso
    especial, igual que el layout distinto de ALMAR WATER.
    """
    enc = [nz.norm_encabezado(v) for v in encabezados_crudos]

    col_nombre = next((i for i, h in enumerate(enc) if h == "NOMBRE"), None)
    inicio_salida = None
    if col_nombre is not None:
        inicio_salida = next(
            (i for i in range(col_nombre + 1, len(enc))
             if enc[i] in config.INICIO_BLOQUE_SALIDA),
            None,
        )

    fin_ingreso = inicio_salida if inicio_salida is not None else len(enc)
    ingreso = _mapear(enc, config.INGRESO_SINONIMOS, 0, fin_ingreso)

    # Rescate del hostal cuando su encabezado esta mal escrito. Del 16 al 17 de
    # julio la celda dice 'p' en vez de 'HOSTAL', y sin esto se perderia el
    # hostal de todas las filas de esas hojas.
    #
    # El ancla es EMPRESA, que si aparece bien escrita: en ambos libros el
    # hostal es la columna inmediatamente anterior. Solo se aplica si esa
    # columna quedo sin asignar, asi que nunca pisa un encabezado valido.
    if "hostal" not in ingreso and "empresa" in ingreso:
        candidata = min(ingreso["empresa"]) - 1
        asignadas = {c for cols in ingreso.values() for c in cols}
        if candidata >= 0 and candidata not in asignadas:
            ingreso["hostal"] = [candidata]

    return Mapeo(
        ingreso=ingreso,
        salida=(_mapear(enc, config.SALIDA_SINONIMOS, inicio_salida, len(enc))
                if inicio_salida is not None else {}),
        inicio_salida=inicio_salida,
    )


def _base(archivo: str, hoja: str, fila_num: int, bloque: str, fecha_hoja) -> dict:
    return {
        "archivo_origen": archivo,
        "hoja": hoja,
        "fila": fila_num,
        "bloque": bloque,
        "fecha_hoja": fecha_hoja,
    }


CAMPOS_INGRESO = (
    "hostal", "empresa", "folio", "habitacion", "tipo_habitacion", "grupo",
    "turno", "nombre", "rut", "celular", "cargo", "motivo", "observaciones",
    "cambio_sabanas", "desayuno", "almuerzo", "cena", "colacion_normal",
    "colacion_especial", "sub",
)

CAMPOS_SALIDA = (
    "hostal", "empresa", "folio", "habitacion", "turno", "grupo", "nombre",
    "rut", "celular", "cargo", "motivo", "observaciones", "chip", "llaves",
    "desayuno", "almuerzo", "cena",
)


def leer_hojas_diarias(ruta: Path, anio: int, mes: int) -> list[dict]:
    """Todas las hojas diarias de un libro, como filas de staging.registro_crudo."""
    wb = load_workbook(ruta, data_only=True, read_only=True)
    archivo = ruta.name
    registros: list[dict] = []

    for hoja in wb.sheetnames:
        dia = nz.dia_desde_hoja(hoja)
        if dia is None:
            continue                      # R. OFICIAL, ALMUERZOS, hojas ocultas
        try:
            fecha_hoja = dt.date(anio, mes, dia)
        except ValueError:
            continue                      # p.ej. 31 en un mes de 30 dias

        ws = wb[hoja]
        filas = list(ws.iter_rows(values_only=True))
        if not filas:
            continue

        mapeo = detectar_mapeo_diario(list(filas[0]))
        perfil = f"salida@{mapeo.inicio_salida}" if mapeo.inicio_salida is not None else "sin-salida"

        for i, fila in enumerate(filas[1:], start=2):
            if fila is None:
                continue

            # Bajo los datos hay un bloque de totales cuya primera columna cae
            # justo en la columna NOMBRE ('HAB', 'TOTAL', 'ISAM'). Al llegar
            # ahi se acaba la hoja: todo lo que sigue son formulas, no personas.
            if nz.es_etiqueta_pie(mapeo.valor(fila, "INGRESO", "nombre")):
                break

            for bloque, campos in (("INGRESO", CAMPOS_INGRESO), ("SALIDA", CAMPOS_SALIDA)):
                nombre = nz.texto(mapeo.valor(fila, bloque, "nombre"))
                if not nombre:
                    continue
                if nombre.replace(".", "").replace(",", "").isdigit():
                    continue
                if nz.es_etiqueta_pie(nombre):
                    continue

                reg = _base(archivo, hoja, i, bloque, fecha_hoja)
                for campo in campos:
                    reg[campo] = nz.texto(mapeo.valor(fila, bloque, campo))
                # La columna CARGO de ISAM se usa a veces para anotar una hora
                # ('16:00:00'), que no es un cargo y no debe entrar al catalogo.
                reg["cargo"] = nz.cargo(mapeo.valor(fila, bloque, "cargo"))
                reg["nombre"] = nombre
                reg["perfil"] = perfil
                reg["fecha"] = mapeo.valor(fila, bloque, "fecha")
                reg["hora"] = mapeo.valor(fila, bloque, "hora")
                registros.append(reg)

    wb.close()
    return registros


def leer_registro_oficial(ruta: Path, anio: int, mes: int) -> list[dict]:
    """Desarma la matriz persona x dia en una fila por noche.

    Esta hoja es la fuente autoritativa de noches: trae la marca D/N/E por dia
    y refleja las estadias con huecos (JUAN CORREA tiene 18 noches en tres
    tramos separados), que no se pueden reconstruir desde ingreso-salida.

    Se ignoran las filas de totales del pie: solo se aceptan filas con nombre.
    """
    wb = load_workbook(ruta, data_only=True, read_only=True)
    archivo = ruta.name
    registros: list[dict] = []

    for hoja in wb.sheetnames:
        h = nz.norm_encabezado(hoja) or ""
        if "OFICIAL" not in h:
            continue
        # 'REGISTRO OFICIAL (1)' es una copia oculta y desactualizada.
        if "(" in hoja or "1)" in h:
            continue

        ws = wb[hoja]
        filas = list(ws.iter_rows(values_only=True))

        # La fila de encabezado es la que contiene 'NOMBRE'; en estas hojas
        # la fila 1 trae las iniciales de los dias de la semana.
        idx_enc = next(
            (i for i, f in enumerate(filas[:5])
             if f and any(nz.norm_encabezado(v) == "NOMBRE" for v in f)),
            None,
        )
        if idx_enc is None:
            continue

        enc = [nz.norm_encabezado(v) for v in filas[idx_enc]]
        campos = _mapear(enc, config.OFICIAL_SINONIMOS, 0, len(enc))

        # Columnas de dia: encabezados tipo '01 JULIO', '2  JULIO'.
        dias: dict[int, int] = {}
        for idx, v in enumerate(filas[idx_enc]):
            d = nz.dia_desde_encabezado(v)
            if d is not None:
                dias[idx] = d
        if not dias:
            continue

        def val(fila, campo):
            for c in campos.get(campo, []):
                if c < len(fila) and fila[c] is not None and str(fila[c]).strip() not in ("", "-"):
                    return fila[c]
            return None

        for i, fila in enumerate(filas[idx_enc + 1:], start=idx_enc + 2):
            if fila is None:
                continue
            nombre = nz.texto(val(fila, "nombre"))
            if not nombre or nombre.replace(".", "").isdigit():
                continue

            comun = {
                "archivo_origen": archivo,
                "hoja": hoja,
                "fila": i,
                "tipo_habitacion": nz.texto(val(fila, "tipo_habitacion")),
                "hostal": nz.texto(val(fila, "hostal")),
                "empresa": nz.texto(val(fila, "empresa")),
                "grupo": nz.texto(val(fila, "grupo")),
                "folio": nz.texto(val(fila, "folio")),
                "habitacion": nz.texto(val(fila, "habitacion")),
                "nombre": nombre,
                "rut": nz.texto(val(fila, "rut")),
                "cargo": nz.cargo(val(fila, "cargo")),
                "observacion": nz.texto(val(fila, "observacion")),
                "almuerzo_extra": nz.texto(val(fila, "almuerzo_extra")),
                "estacionamiento": nz.texto(val(fila, "estacionamiento")),
                "patente": nz.texto(val(fila, "patente")),
                "total_alojamiento": nz.texto(val(fila, "total_alojamiento")),
                "fecha_ingreso": nz.fecha(val(fila, "fecha_ingreso")),
            }

            for col, dia in dias.items():
                if col >= len(fila):
                    continue
                marca = nz.texto(fila[col])
                if not marca:
                    continue
                try:
                    f = dt.date(anio, mes, dia)
                except ValueError:
                    continue
                registros.append({**comun, "fecha": f, "marca": marca})

    wb.close()
    return registros


def leer_almuerzos(ruta: Path, anio: int, mes: int) -> list[dict]:
    """Hoja ALMUERZOS ISAM: un servicio por fila.

    Las filas 'SIN ALMUERZOS' son marcadores de "ese dia no hubo"; se saltan,
    porque crear una persona con ese nombre ensuciaria el catalogo. La ausencia
    de almuerzos se deriva por ausencia de filas.
    """
    wb = load_workbook(ruta, data_only=True, read_only=True)
    archivo = ruta.name
    registros: list[dict] = []

    for hoja in wb.sheetnames:
        h = nz.norm_encabezado(hoja) or ""
        if "ALMUERZO" not in h:
            continue

        ws = wb[hoja]
        filas = list(ws.iter_rows(values_only=True))
        idx_enc = next(
            (i for i, f in enumerate(filas[:10])
             if f and any(nz.norm_encabezado(v) == "NOMBRE" for v in f)),
            None,
        )
        if idx_enc is None:
            continue

        enc = [nz.norm_encabezado(v) for v in filas[idx_enc]]
        campos = _mapear(enc, config.ALMUERZO_SINONIMOS, 0, len(enc))
        # La columna de autorizacion ('SRA. ELIANA') va sin encabezado,
        # inmediatamente despues de la ultima columna mapeada.
        ultima = max((c for cols in campos.values() for c in cols), default=-1)
        col_autoriza = ultima + 1

        for i, fila in enumerate(filas[idx_enc + 1:], start=idx_enc + 2):
            if fila is None:
                continue

            def val(campo):
                for c in campos.get(campo, []):
                    if c < len(fila) and fila[c] is not None and str(fila[c]).strip() not in ("", "-"):
                        return fila[c]
                return None

            nombre = nz.texto(val("nombre"))
            if not nombre or nz.es_marcador_sin_datos(nombre):
                continue

            registros.append({
                "archivo_origen": archivo,
                "hoja": hoja,
                "fila": i,
                "fecha": nz.fecha(val("fecha")),
                "nombre": nombre,
                "rut": nz.texto(val("rut")),
                "empresa": nz.texto(val("empresa")),
                "tipo_servicio": nz.texto(val("tipo_servicio")),
                "cantidad": nz.cantidad(val("cantidad")),
                "hostal": nz.texto(val("hostal")),
                "autorizado_por": (nz.texto(fila[col_autoriza])
                                   if col_autoriza < len(fila) else None),
            })

    wb.close()
    return registros
